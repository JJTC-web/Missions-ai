"""
Needs Assessment Workbook generator.

generate_workbook_content() calls Claude for the four analytical sections
(target areas, volunteer needs, budget, logic model). build_workbook_xlsx()
writes the full 7-tab .xlsx matching the reference "MissionOS Community
Needs Explorer" workbook's structure, styling, and formulas, then
recalculates it so cached formula values are correct.
"""

import json
from pathlib import Path

import anthropic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import needs_prompt

MODEL = "claude-opus-5"

FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, bold=True, color="FF1F3864", size=14)
SUBTITLE_FONT = Font(name=FONT_NAME, color="FF595959")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
ALT_FILL = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
BODY_FONT = Font(name=FONT_NAME)
INPUT_FONT = Font(name=FONT_NAME, color="FF0000FF")
BOLD_FONT = Font(name=FONT_NAME, bold=True)
LABEL_FONT = Font(name=FONT_NAME, bold=True)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

CURRENCY_FORMAT = "\\$#,##0"
PERCENT_FORMAT = "0.0%"


def generate_workbook_content(org, region, region_stats, resource_directory):
    """Calls Claude for the four analytical sections. Raises on API/network failure."""
    prompt = needs_prompt.build_user_prompt(org, region, region_stats, resource_directory)
    client = anthropic.Anthropic()
    response = client.messages.create(
        model=MODEL,
        max_tokens=12000,
        system=needs_prompt.SYSTEM_PROMPT,
        output_config={"format": {"type": "json_schema", "schema": needs_prompt.NEEDS_WORKBOOK_SCHEMA}},
        messages=[{"role": "user", "content": prompt}],
    )
    text = next(block.text for block in response.content if block.type == "text")
    return json.loads(text)


def _new_sheet(wb, name, title, subtitle, first_sheet=False):
    ws = wb.active if first_sheet else wb.create_sheet(name)
    if first_sheet:
        ws.title = name
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    return ws


def _write_header(ws, row, headers):
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate


def _apply_row_fill(ws, row, num_cols, first_data_row):
    if (row - first_data_row) % 2 == 1:
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = ALT_FILL


def _set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_how_to_use(wb, region_label):
    ws = _new_sheet(
        wb,
        "How to Use",
        "MissionOS — Community Needs Explorer",
        f"Generated for: {region_label} — built from data on file in MissionOS",
        first_sheet=True,
    )
    _set_widths(ws, {"A": 22, "B": 70})
    rows = [
        ("Purpose", "This workbook is auto-generated when a user clicks a city/region in MissionOS. It turns public need data + your local resource directory into a prioritized action plan."),
        ("Tab: Regional Snapshot", "The raw need data pulled for the clicked region — homelessness, food insecurity, poverty — with sources and caveats. This is the evidence base everything else points back to."),
        ("Tab: Local Resource Directory", "Existing service providers in the region (shelters, pantries, etc.) — what capacity already exists, so MissionOS doesn't recommend duplicating it."),
        ("Tab: Target Areas of Support", "Prioritization framework: scores each need category by severity x local capacity gap, ranks where a new program or partner should focus first."),
        ("Tab: Volunteer Needs Plan", "Translates each priority area into roles, hours, and headcount using standard volunteer-program ratios."),
        ("Tab: Budget Capture", "Line-item budget by category, checked against the standard nonprofit 65/35 program-to-overhead benchmark."),
        ("Tab: Success Framework", "A Logic Model (Inputs -> Activities -> Outputs -> Outcomes -> Impact) with KPIs and measurement cadence for each priority area, so 'success' is defined before work starts."),
        ("Color legend", "Blue text = hardcoded/auto-populated input. Black text = formula, recalculates automatically."),
        ("Data caveat", "City-level PIT (Point-in-Time) homeless counts are rarely published separately from county totals. Where MissionOS only has a county figure, it is labeled as such — never presented as city-specific."),
    ]
    row = 4
    for label, text in rows:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=row, column=2, value=text)
        cell.font = BODY_FONT
        cell.alignment = WRAP_TOP
        ws.row_dimensions[row].height = 31.5
        row += 1


def _write_regional_snapshot(wb, region, region_stats):
    ws = _new_sheet(
        wb,
        "Regional Snapshot",
        "Regional Snapshot",
        "Auto-populated when a region is clicked; MissionOS pulls the latest figures available for that geography",
    )
    _set_widths(ws, {"A": 34, "B": 20, "C": 18, "D": 55})
    _write_header(ws, 4, ["Metric", "Value", "Geography Level", "Source / Caveat"])

    rows = [("Region clicked", f"{region['city']}, {region['state']}", "City", "User selection")]
    if region.get("county"):
        rows.append(("County", f"{region['county']} County, {region['state']}", "County", "—"))
    for stat in region_stats:
        rows.append((stat["metric_name"], stat["value"], stat.get("geography_level") or "—", stat.get("source") or "—"))

    first_data_row = 5
    for i, (metric, value, level, source) in enumerate(rows):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=metric).font = BODY_FONT
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.font = INPUT_FONT
        ws.cell(row=row, column=3, value=level).font = BODY_FONT
        ws.cell(row=row, column=4, value=source).font = BODY_FONT
        _apply_row_fill(ws, row, 4, first_data_row)


def _write_resource_directory(wb, region, resource_directory):
    ws = _new_sheet(
        wb,
        "Local Resource Directory",
        "Local Resource Directory",
        "Existing providers MissionOS found for this region — the current capacity baseline",
    )
    _set_widths(ws, {"A": 32, "B": 26, "C": 24, "D": 20, "E": 15, "F": 10})
    _write_header(ws, 4, ["Organization", "Address", "Services", "Population Served", "Phone", "County"])

    first_data_row = 5
    for i, entry in enumerate(resource_directory):
        row = first_data_row + i
        values = [
            entry["name"],
            entry.get("address") or "—",
            entry.get("services") or "—",
            entry.get("population_served") or "—",
            entry.get("phone") or "—",
            region.get("county") or "—",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value).font = BODY_FONT
        _apply_row_fill(ws, row, 6, first_data_row)


def _write_target_areas(wb, target_areas):
    ws = _new_sheet(
        wb,
        "Target Areas of Support",
        "Target Areas of Support",
        "Framework: Severity x Capacity Gap prioritization (adapted from HUD Continuum of Care gap analysis)",
    )
    _set_widths(ws, {"A": 30, "B": 34, "C": 12, "D": 14, "E": 12, "F": 12, "G": 34, "H": 26})
    _write_header(ws, 4, ["Need Category", "Existing Local Capacity", "Severity (1-5)", "Capacity Gap (1-5)", "Priority Score", "Priority Rank", "Recommended Action", "Framework Reference"])

    first_data_row = 5
    last_data_row = first_data_row + len(target_areas) - 1
    rank_range = f"$E${first_data_row}:$E${last_data_row}"

    for i, area in enumerate(target_areas):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=area["category"]).font = BODY_FONT
        ws.cell(row=row, column=2, value=area["existingCapacity"]).font = BODY_FONT
        ws.cell(row=row, column=3, value=area["severity"]).font = INPUT_FONT
        ws.cell(row=row, column=4, value=area["capacityGap"]).font = INPUT_FONT
        ws.cell(row=row, column=5, value=f"=C{row}*D{row}").font = BODY_FONT
        ws.cell(row=row, column=6, value=f"=RANK(E{row},{rank_range})").font = BODY_FONT
        ws.cell(row=row, column=7, value=area["recommendedAction"]).font = BODY_FONT
        ws.cell(row=row, column=8, value=area["framework"]).font = BODY_FONT
        _apply_row_fill(ws, row, 8, first_data_row)


def _write_volunteer_needs(wb, volunteer_needs):
    ws = _new_sheet(
        wb,
        "Volunteer Needs Plan",
        "Volunteer Needs Plan",
        "Framework: role-to-hours ratios adapted from AmeriCorps/VolunteerMatch program benchmarks",
    )
    _set_widths(ws, {"A": 30, "B": 26, "C": 30, "D": 16, "E": 14, "F": 16, "G": 26})
    _write_header(ws, 4, ["Priority Area", "Volunteer Role", "Skills Needed", "Hrs / Volunteer / Month", "Volunteers Needed", "Total Monthly Hours", "Recruitment Channel"])

    first_data_row = 5
    for i, need in enumerate(volunteer_needs):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=need["priorityArea"]).font = BODY_FONT
        ws.cell(row=row, column=2, value=need["role"]).font = BODY_FONT
        ws.cell(row=row, column=3, value=need["skills"]).font = BODY_FONT
        ws.cell(row=row, column=4, value=need["hoursPerVolunteer"]).font = INPUT_FONT
        ws.cell(row=row, column=5, value=need["volunteersNeeded"]).font = INPUT_FONT
        ws.cell(row=row, column=6, value=f"=D{row}*E{row}").font = BODY_FONT
        ws.cell(row=row, column=7, value=need["channel"]).font = BODY_FONT
        _apply_row_fill(ws, row, 7, first_data_row)

    total_row = first_data_row + len(volunteer_needs)
    last_data_row = total_row - 1
    ws.cell(row=total_row, column=5, value="Total").font = BOLD_FONT
    ws.cell(row=total_row, column=6, value=f"=SUM(F{first_data_row}:F{last_data_row})").font = BOLD_FONT


def _write_budget(wb, budget_items):
    ws = _new_sheet(
        wb,
        "Budget Capture",
        "Budget Capture",
        "Framework: standard nonprofit 65/35 program-to-overhead benchmark (Charity Navigator / BBB Wise Giving Alliance)",
    )
    _set_widths(ws, {"A": 14, "B": 40, "C": 30, "D": 16, "E": 26, "F": 32})
    _write_header(ws, 4, ["Category", "Line Item", "Priority Area", "Est. Annual Cost", "Target Funding Source", "Notes"])

    first_data_row = 5
    for i, item in enumerate(budget_items):
        row = first_data_row + i
        ws.cell(row=row, column=1, value=item["category"]).font = BODY_FONT
        ws.cell(row=row, column=2, value=item["lineItem"]).font = BODY_FONT
        ws.cell(row=row, column=3, value=item["priorityArea"]).font = BODY_FONT
        cost_cell = ws.cell(row=row, column=4, value=item["estCost"])
        cost_cell.font = INPUT_FONT
        cost_cell.number_format = CURRENCY_FORMAT
        ws.cell(row=row, column=5, value=item["fundingSource"]).font = BODY_FONT
        ws.cell(row=row, column=6, value=item["notes"]).font = BODY_FONT
        _apply_row_fill(ws, row, 6, first_data_row)

    last_data_row = first_data_row + len(budget_items) - 1
    total_row = last_data_row + 2
    program_row = total_row + 1
    pct_row = total_row + 2

    ws.cell(row=total_row, column=1, value="Total budget").font = BOLD_FONT
    total_cell = ws.cell(row=total_row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})")
    total_cell.font = BOLD_FONT
    total_cell.number_format = CURRENCY_FORMAT

    ws.cell(row=program_row, column=1, value="Program spend").font = BODY_FONT
    program_cell = ws.cell(row=program_row, column=4, value=f'=SUMIF(A{first_data_row}:A{last_data_row},"Program",D{first_data_row}:D{last_data_row})')
    program_cell.font = BODY_FONT
    program_cell.number_format = CURRENCY_FORMAT

    ws.cell(row=pct_row, column=1, value="Program % of total (target >= 65%)").font = BOLD_FONT
    pct_cell = ws.cell(row=pct_row, column=4, value=f"=D{program_row}/D{total_row}")
    pct_cell.font = BOLD_FONT
    pct_cell.number_format = PERCENT_FORMAT


def _write_success_framework(wb, logic_model):
    ws = _new_sheet(
        wb,
        "Success Framework",
        "Success Framework",
        "Logic Model: Inputs -> Activities -> Outputs -> Outcomes -> Impact, with KPIs per priority area",
    )
    _set_widths(ws, {"A": 26, "B": 24, "C": 26, "D": 22, "E": 30, "F": 26, "G": 18})
    _write_header(ws, 4, ["Priority Area", "Inputs", "Activities", "Outputs (count)", "Outcomes (change)", "Impact (long-term)", "Measurement Cadence"])

    first_data_row = 5
    for i, entry in enumerate(logic_model):
        row = first_data_row + i
        values = [
            entry["priorityArea"], entry["inputs"], entry["activities"],
            entry["outputs"], entry["outcomes"], entry["impact"], entry["cadence"],
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value).font = BODY_FONT
        _apply_row_fill(ws, row, 7, first_data_row)


def build_workbook_xlsx(org, region, region_stats, resource_directory, content, output_path):
    """Writes the full 7-tab workbook to output_path.

    Formula cells (Priority Score/Rank, volunteer/budget totals, the 65/35
    program-spend check) are written as live Excel formulas, not
    precomputed values -- Excel/Sheets/LibreOffice recalculate them
    automatically the moment the file is opened, so no server-side
    recalculation step is needed (or appropriate: this app doesn't ship
    LibreOffice, and shelling out to it per-request wouldn't be a sound
    dependency for a web server even if it did).
    """
    region_label = f"{region['city']}, {region['state']}"
    if region.get("county"):
        region_label += f" ({region['county']} County)"

    wb = Workbook()
    _write_how_to_use(wb, region_label)
    _write_regional_snapshot(wb, region, region_stats)
    _write_resource_directory(wb, region, resource_directory)
    _write_target_areas(wb, content["targetAreas"])
    _write_volunteer_needs(wb, content["volunteerNeeds"])
    _write_budget(wb, content["budgetItems"])
    _write_success_framework(wb, content["logicModel"])

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
